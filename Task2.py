from openpyxl import Workbook
from datetime import datetime, timedelta

# Function to get the name of the month for a given date
def get_month_name(date):
    return date.strftime("%b")

# Get the current date
current_date = datetime.now()

# Get the names of the current month, previous month, and the month before that
current_month = get_month_name(current_date) # will get current month
previous_month = get_month_name(current_date - timedelta(days=current_date.day)) # will get current month -1
month_before_previous = get_month_name(current_date - timedelta(days=current_date.day + 30))  # will get current month -2

# Create a new workbook and add sheets
wb = Workbook()

# Add sheets for the current month, previous month, and the month before the previous month
wb.create_sheet(current_month)
wb.create_sheet(previous_month)
wb.create_sheet(month_before_previous)

# Remove the default sheet created by openpyxl
default_sheet = wb["Sheet"]
wb.remove(default_sheet)

# Save the workbook with the name 'niki.xlsx'
wb.save("niki.xlsx")

print("Excel file 'niki.xlsx' created with sheets:", current_month, previous_month, month_before_previous)
