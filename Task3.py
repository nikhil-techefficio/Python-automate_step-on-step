#Now we have an xl file by name dummy which has jun jul may need to read the data and past the same in our newly created sheet

from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta

# Function to get the name of the month for a given date
def get_month_name(date):
    return date.strftime("%b")

# Get the current date
current_date = datetime.now()

# Get the names of the current month, previous month, and the month before that
current_month = get_month_name(current_date)
previous_month = get_month_name(current_date - timedelta(days=current_date.day))
month_before_previous = get_month_name(current_date - timedelta(days=current_date.day + 30))

# Load the existing workbook 'dummy.xlsx'
dummy_wb = load_workbook('dummy.xlsx')

# Create a new workbook and add sheets
new_wb = Workbook()

# Add sheets for the current month, previous month, and the month before the previous month
new_wb.create_sheet(current_month)
new_wb.create_sheet(previous_month)
new_wb.create_sheet(month_before_previous)

# Remove the default sheet created by openpyxl
default_sheet = new_wb["Sheet"]
new_wb.remove(default_sheet)

# Function to copy data from one sheet to another
def copy_sheet_data(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value

# Copy data from 'dummy.xlsx' sheets to the new workbook's sheets
if current_month in dummy_wb.sheetnames:
    copy_sheet_data(dummy_wb[current_month], new_wb[current_month])
if previous_month in dummy_wb.sheetnames:
    copy_sheet_data(dummy_wb[previous_month], new_wb[previous_month])
if month_before_previous in dummy_wb.sheetnames:
    copy_sheet_data(dummy_wb[month_before_previous], new_wb[month_before_previous])

# Save the new workbook with the name 'niki.xlsx'
new_wb.save('niki.xlsx')

print(f"Data copied to 'niki.xlsx' for sheets: {current_month}, {previous_month}, {month_before_previous}")
